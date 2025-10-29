import os
import logging
import pypdf
import json
import csv
import openpyxl  # 用于 Excel
import mimetypes # 用于检测文件类型

# 附件文件夹
ATTACHMENTS_DIR = "benchmark/attachments/"

logger = logging.getLogger(__name__)

def read_file(filename: str) -> str:
    """
    Reads the text content of various file types (.txt, .pdf, .json, .csv, .xlsx) 
    located in the benchmark/attachments/ directory.

    If the file is an image, it will return an error message telling 
    the agent to use a different tool.

    Args:
        filename: The name of the file to read (e.g., "summary.pdf", "data.json").

    Returns:
        The text content of the file, or an error message if it fails.
    """
    
    # 构建安全的文件路径
    file_path = os.path.join(ATTACHMENTS_DIR, filename)

    # 1. 检查文件是否存在
    if not os.path.exists(file_path):
        return f"Error: File not found at path: {file_path}. Please ensure the filename is correct and it is in the attachments folder."

    # 2. 检查文件是否为图像 (MIME type)
    mimetype, _ = mimetypes.guess_type(file_path)
    if mimetype and mimetype.startswith('image/'):
        return (
            f"Error: The file '{filename}' is an image. "
            "This tool can only read text-based files. "
            "Please use an image analysis tool to read this file."
        )

    try:
        # --- 处理 PDF 文件 ---
        if filename.endswith('.pdf'):
            content = ""
            with open(file_path, 'rb') as f:
                reader = pypdf.PdfReader(f)
                if not reader.pages:
                    return f"Error: PDF file '{filename}' is empty or corrupted."
                for page in reader.pages:
                    page_text = page.extract_text()
                    if page_text:
                        content += page_text + "\n"
            return content

        # --- 处理 TXT 文件 ---
        elif filename.endswith('.txt'):
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        
        # --- (新) 处理 JSON 文件 ---
        elif filename.endswith('.json'):
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                # 将 JSON 格式化为易于 LLM 阅读的字符串
                return json.dumps(data, indent=2)

        # --- (新) 处理 CSV 文件 ---
        elif filename.endswith('.csv'):
            content = ""
            with open(file_path, 'r', encoding='utf-8', newline='') as f:
                reader = csv.reader(f)
                for row in reader:
                    # 将每行转换为逗号分隔的字符串
                    content += ",".join(row) + "\n"
            return content

        # --- (新) 处理 Excel 文件 ---
        elif filename.endswith('.xlsx'):
            content = ""
            workbook = openpyxl.load_workbook(file_path, data_only=True) # data_only=True 转换公式为值
            for sheet_name in workbook.sheetnames:
                content += f"--- Sheet: {sheet_name} ---\n"
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    row_data = [str(cell.value) if cell.value is not None else "" for cell in row]
                    content += ",".join(row_data) + "\n"
                content += "\n"
            return content

        else:
            return (
                f"Error: Unsupported file type for: {filename}. "
                "This tool can currently read: .txt, .pdf, .json, .csv, .xlsx. "
                f"The file '{filename}' is not one of these types."
            )

    except Exception as e:
        return f"Error while reading file '{filename}': {str(e)}"